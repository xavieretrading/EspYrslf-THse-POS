import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define paths
scratch_dir = r"c:\Users\Philippines Freight\.gemini\antigravity-ide\brain\4f7b6118-c5c9-4ac0-bad1-30e86099146c\scratch"
rec_dir = r"C:\Users\Philippines Freight\MainSystems\POS\recordsExcel"
output_file = os.path.join(rec_dir, "S1P_and_Sp1n_Inventory_Stocks_Report.xlsx")

# Load data
with open(os.path.join(scratch_dir, "categories.json"), "r") as f:
    categories = json.load(f)

with open(os.path.join(scratch_dir, "products_branch_27.json"), "r") as f:
    products = json.load(f)

with open(os.path.join(scratch_dir, "transactions_branch_27.json"), "r") as f:
    transactions = json.load(f)

cat_map = {c["id"]: c for c in categories}

# Group transactions by product ID
tx_by_product = {}
for t in transactions:
    pid = t["product_id"]
    if pid not in tx_by_product:
        tx_by_product[pid] = []
    tx_by_product[pid].append(t)

laundry_rows = []
retail_rows = []

# Helper to format date
def format_db_date(date_str):
    try:
        day_part = date_str[:10]
        dt = datetime.strptime(day_part, "%Y-%m-%d")
        return dt.strftime("%B %d, %Y")
    except Exception:
        return date_str

for p in products:
    cat = cat_map.get(p["category_id"])
    if not cat:
        continue
    
    division = cat.get("division", "coffee")
    
    # Process transactions to find beginning stock and additions
    p_txs = tx_by_product.get(p["id"], [])
    in_txs = [t for t in p_txs if t["type"] == "in"]
    in_txs.sort(key=lambda x: x["created_at"])
    
    beg_stock = 0
    add_qty = None
    add_date = None
    date_inserted = "July 31, 2026"  # Default to initial setup
    
    if len(in_txs) > 0:
        first_tx = in_txs[0]
        # First 'in' is beginning stock
        beg_stock = first_tx["quantity"]
        
        # Calculate Date Added (when it was inserted/added to system)
        created_time = first_tx["created_at"]
        if "Initial stock setup" in first_tx.get("remarks", "") and created_time <= "2026-08-01T23:59:59":
            date_inserted = "July 31, 2026"
        else:
            date_inserted = format_db_date(created_time)
            
        # Subsequent 'in' transactions with quantity > 0 are additions
        additions = [t for t in in_txs[1:] if t["quantity"] > 0]
        if additions:
            add_qty = sum(t["quantity"] for t in additions)
            add_date = ", ".join(format_db_date(t["created_at"]) for t in additions)
    else:
        # No 'in' transactions, skip if inactive with 0 stock
        if p["is_active"] == 0 and p["stock"] == 0:
            continue
        beg_stock = p["stock"]
        date_inserted = "July 31, 2026"
    
    # Clean up name if it ends with colon
    name = p["name"].strip()
    if name.endswith(":"):
        name = name[:-1].strip()
        
    row = {
        "date_added": date_inserted,
        "item": name,
        "beginning_stock": beg_stock,
        "additional_stock": add_qty,
        "date_additional": add_date
    }
    
    if division == "laundry":
        laundry_rows.append(row)
    else:
        retail_rows.append(row)

# Sort rows alphabetically by item name
laundry_rows.sort(key=lambda x: x["item"])
retail_rows.sort(key=lambda x: x["item"])

# Create Workbook
wb = openpyxl.Workbook()
# Remove default sheet
wb.remove(wb.active)

# Styles
font_name = "Segoe UI"
title_font = Font(name=font_name, size=16, bold=True)
subtitle_font = Font(name=font_name, size=10, italic=True, color="595959")
header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
data_font = Font(name=font_name, size=11)
bold_data_font = Font(name=font_name, size=11, bold=True)

thin_border_side = Side(border_style="thin", color="D3D3D3")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

laundry_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Classic Navy
retail_header_fill = PatternFill(start_color="70543E", end_color="70543E", fill_type="solid") # Warm Coffee Brown

zebra_laundry_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid") # Light Blue-Grey
zebra_retail_fill = PatternFill(start_color="FAF6F0", end_color="FAF6F0", fill_type="solid") # Light Cream/Brown

highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green for additions

# Headers definition
headers = ["Date Added", "Item Description", "Beginning Stock", "Additional Stock", "Date Added (Additional)"]

def build_sheet(ws, title, subtitle, rows, header_fill, zebra_fill):
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.cell(row=2, column=1, value=title).font = title_font
    ws.cell(row=3, column=1, value=subtitle).font = subtitle_font
    
    # Table Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Data Rows
    current_row = 6
    for idx, r in enumerate(rows):
        # Determine cell values
        date_added = r["date_added"]
        item = r["item"]
        beg_stock = r["beginning_stock"]
        add_stock = r["additional_stock"] if r["additional_stock"] is not None else ""
        date_additional = r["date_additional"] if r["date_additional"] else ""
        
        # Format beginning stock if it is a service placeholder
        if beg_stock in [9999, 9998]:
            beg_stock_val = "Unlimited"
        else:
            beg_stock_val = beg_stock
            
        c_date = ws.cell(row=current_row, column=1, value=date_added)
        c_item = ws.cell(row=current_row, column=2, value=item)
        c_beg = ws.cell(row=current_row, column=3, value=beg_stock_val)
        c_add = ws.cell(row=current_row, column=4, value=add_stock)
        c_add_date = ws.cell(row=current_row, column=5, value=date_additional)
        
        # Alignments
        c_date.alignment = Alignment(horizontal="center", vertical="center")
        c_item.alignment = Alignment(horizontal="left", vertical="center")
        c_beg.alignment = Alignment(horizontal="center", vertical="center")
        c_add.alignment = Alignment(horizontal="center", vertical="center")
        c_add_date.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set fonts
        for cell in [c_date, c_item, c_beg, c_add, c_add_date]:
            cell.font = data_font
            cell.border = thin_border
            
        # Alternating zebra fills
        if idx % 2 == 1:
            c_date.fill = zebra_fill
            c_item.fill = zebra_fill
            c_beg.fill = zebra_fill
            
        # Highlight additions in soft green
        if add_stock != "":
            c_add.fill = highlight_fill
            c_add.font = bold_data_font
            c_add_date.fill = highlight_fill
            c_add_date.font = bold_data_font
        elif idx % 2 == 1:
            c_add.fill = zebra_fill
            c_add_date.fill = zebra_fill
            
        current_row += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Skip title rows for width calculation
        for cell in col[4:]: 
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
                
        # Handle headers length
        max_len = max(max_len, len(str(col[4].value or "")))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Set height for header row
    ws.row_dimensions[5].height = 25
    
    # Set height for data rows
    for r in range(6, current_row):
        ws.row_dimensions[r].height = 20

# Create Sheet 1: Laundry Inventory
ws1 = wb.create_sheet(title="Laundry Inventory")
build_sheet(
    ws1, 
    "S1P & SP1N LAUNDRY SHOP - LAUNDRY INVENTORY", 
    "System entry dates, beginning stocks, and subsequent stock additions for services and detergents", 
    laundry_rows, 
    laundry_header_fill, 
    zebra_laundry_fill
)

# Create Sheet 2: Retail & Cafe Inventory
ws2 = wb.create_sheet(title="Retail & Cafe Inventory")
build_sheet(
    ws2, 
    "S1P & SP1N LAUNDRY SHOP - RETAIL & CAFE INVENTORY", 
    "System entry dates, beginning stocks, and subsequent stock additions for beverages, snacks and desserts", 
    retail_rows, 
    retail_header_fill, 
    zebra_retail_fill
)

# Save the workbook with fallback
try:
    wb.save(output_file)
    print(f"SUCCESS: Excel file saved at: {output_file}")
except PermissionError:
    fallback_file = os.path.join(rec_dir, "S1P_and_Sp1n_Inventory_Stocks_Report_v2.xlsx")
    wb.save(fallback_file)
    print(f"FALLBACK: Main file was locked. Excel file saved at: {fallback_file}")
